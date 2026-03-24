from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime
import os, shutil, uuid, json

from ...core.database import get_db
from ...core.security import get_current_user
from ...core.config import settings
from ...models.project import Project, ProjectDocument, ProjectNote, ProjectStatus, ProjectType
from ...models.user import User

router = APIRouter()


# ── Schemas ──────────────────────────────────────────────────────────────
class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    status: ProjectStatus = ProjectStatus.draft
    project_type: ProjectType = ProjectType.equity
    target_raise: Optional[float] = None
    minimum_investment: Optional[float] = None
    currency: str = "USD"
    irr_target: Optional[float] = None
    equity_multiple_target: Optional[float] = None
    hold_period_years: Optional[float] = None
    country: Optional[str] = None
    sector: Optional[str] = None
    sub_sector: Optional[str] = None
    deal_open_date: Optional[datetime] = None
    deal_close_date: Optional[datetime] = None
    expected_close_date: Optional[datetime] = None
    sponsor_name: Optional[str] = None
    sponsor_contact: Optional[str] = None
    source: Optional[str] = None
    tags: Optional[List[str]] = []
    custom_fields: Optional[dict] = {}


class ProjectUpdate(ProjectCreate):
    name: Optional[str] = None


def _auto_code(db: Session) -> str:
    count = db.query(Project).count() + 1
    return f"PRJ-{count:04d}"


@router.get("/")
def list_projects(
    skip: int = 0,
    limit: int = Query(50, le=200),
    status: Optional[ProjectStatus] = None,
    project_type: Optional[ProjectType] = None,
    search: Optional[str] = None,
    country: Optional[str] = None,
    sector: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Project)
    if status:
        q = q.filter(Project.status == status)
    if project_type:
        q = q.filter(Project.project_type == project_type)
    if country:
        q = q.filter(Project.country.ilike(f"%{country}%"))
    if sector:
        q = q.filter(Project.sector.ilike(f"%{sector}%"))
    if search:
        q = q.filter(
            (Project.name.ilike(f"%{search}%")) |
            (Project.code.ilike(f"%{search}%")) |
            (Project.description.ilike(f"%{search}%"))
        )
    total = q.count()
    projects = q.order_by(Project.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": projects}


@router.post("/", status_code=201)
def create_project(
    req: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = Project(
        **req.model_dump(),
        code=_auto_code(db),
        created_by_id=current_user.id,
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@router.get("/stats")
def project_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    total = db.query(Project).count()
    by_status = {}
    for s in ProjectStatus:
        by_status[s.value] = db.query(Project).filter(Project.status == s).count()
    by_type = {}
    for t in ProjectType:
        by_type[t.value] = db.query(Project).filter(Project.project_type == t).count()
    return {"total": total, "by_status": by_status, "by_type": by_type}


@router.get("/{project_id}")
def get_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@router.put("/{project_id}")
def update_project(
    project_id: int,
    req: ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    for field, value in req.model_dump(exclude_unset=True).items():
        setattr(project, field, value)
    db.commit()
    db.refresh(project)
    return project


@router.delete("/{project_id}", status_code=204)
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "analyst"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(project)
    db.commit()


# ── Documents ─────────────────────────────────────────────────────────────
@router.post("/{project_id}/documents", status_code=201)
async def upload_document(
    project_id: int,
    file: UploadFile = File(...),
    doc_category: str = Form("general"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    upload_dir = os.path.join(settings.UPLOAD_DIR, "projects", str(project_id))
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename)[1]
    stored_name = f"{uuid.uuid4()}{ext}"
    storage_path = os.path.join(upload_dir, stored_name)

    with open(storage_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    doc = ProjectDocument(
        project_id=project_id,
        filename=stored_name,
        original_filename=file.filename,
        file_type=file.content_type,
        storage_path=storage_path,
        doc_category=doc_category,
        uploaded_by_id=current_user.id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return doc


@router.get("/{project_id}/documents")
def list_documents(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return db.query(ProjectDocument).filter(ProjectDocument.project_id == project_id).all()


@router.delete("/{project_id}/documents/{doc_id}", status_code=204)
def delete_document(
    project_id: int,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.query(ProjectDocument).filter(
        ProjectDocument.id == doc_id,
        ProjectDocument.project_id == project_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    try:
        os.remove(doc.storage_path)
    except FileNotFoundError:
        pass
    db.delete(doc)
    db.commit()


# ── Notes ─────────────────────────────────────────────────────────────────
class NoteCreate(BaseModel):
    content: str
    note_type: str = "general"


@router.post("/{project_id}/notes", status_code=201)
def add_note(
    project_id: int,
    req: NoteCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    note = ProjectNote(
        project_id=project_id,
        content=req.content,
        note_type=req.note_type,
        author_id=current_user.id,
    )
    db.add(note)
    db.commit()
    db.refresh(note)
    return note


@router.get("/{project_id}/notes")
def list_notes(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return db.query(ProjectNote).filter(ProjectNote.project_id == project_id).all()


# ── Excel Upload ──────────────────────────────────────────────────────────
@router.post("/upload-excel", status_code=201)
async def upload_projects_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk-create projects from an Excel (.xlsx) or CSV file.

    Expected columns (case-insensitive):
    name, description, status, project_type, target_raise, minimum_investment,
    currency, irr_target, equity_multiple_target, hold_period_years,
    country, sector, sub_sector, sponsor_name, sponsor_contact, source, tags
    """
    import io
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls") or file.filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted")

    content = await file.read()
    created = []
    errors = []

    try:
        import openpyxl
        from io import BytesIO

        if file.filename.endswith(".csv"):
            import csv, io as _io
            reader = csv.DictReader(_io.StringIO(content.decode("utf-8-sig")))
            rows = list(reader)
        else:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})

        valid_statuses = {s.value for s in ProjectStatus}
        valid_types = {t.value for t in ProjectType}

        for i, row in enumerate(rows, start=2):
            name = row.get("name", "").strip()
            if not name:
                errors.append({"row": i, "error": "Missing 'name'"})
                continue
            try:
                raw_status = row.get("status", "draft").strip().lower()
                status = raw_status if raw_status in valid_statuses else "draft"
                raw_type = row.get("project_type", "equity").strip().lower()
                ptype = raw_type if raw_type in valid_types else "equity"

                def _float(v):
                    try:
                        return float(v) if v else None
                    except Exception:
                        return None

                tags_raw = row.get("tags", "")
                tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

                project = Project(
                    name=name,
                    code=_auto_code(db),
                    description=row.get("description") or None,
                    status=status,
                    project_type=ptype,
                    target_raise=_float(row.get("target_raise")),
                    minimum_investment=_float(row.get("minimum_investment")),
                    currency=row.get("currency", "USD") or "USD",
                    irr_target=_float(row.get("irr_target")),
                    equity_multiple_target=_float(row.get("equity_multiple_target")),
                    hold_period_years=_float(row.get("hold_period_years")),
                    country=row.get("country") or None,
                    sector=row.get("sector") or None,
                    sub_sector=row.get("sub_sector") or None,
                    sponsor_name=row.get("sponsor_name") or None,
                    sponsor_contact=row.get("sponsor_contact") or None,
                    source=row.get("source") or None,
                    tags=tags,
                    created_by_id=current_user.id,
                )
                db.add(project)
                db.flush()
                created.append({"row": i, "name": name, "code": project.code, "id": project.id})
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        db.commit()
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Add 'openpyxl' to requirements.")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    return {"created": len(created), "errors": len(errors), "projects": created, "error_details": errors}
